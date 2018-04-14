import openpyxl

workbook = openpyxl.load_workbook('cwd_file.xlxs')

# a sinle xlxs file might contain multiple tabs and 
# or maybe you can call that a sheet
# that little tab thingy in the left-bottom of your 
# excel file??? yeah, that's what it is... that's a sheet

# you can select a sheet by

sheet1 = workbook.get_sheet_by_name('sheet_name_i.e_sheet1')

# get names of all sheets, no paras

sheets = workbook.get_sheet_names()
# print (sheets) will result in sheet1

# let's suppose we get a sheet named sheet1, let's work with it
cellOrColumnName = sheet1['firstColumnName']
# [] bracket in sheet gets call objects
# cell objects have value member variable
# with all the contents of that cell. 

store = str(cellOrColumnName.value)
#  cell() method retusn a Cell object from a sheet


# EDITING EXCEL SPREADSHEETS
wb = openpyxl.workbook()
# create a workbook object
wb.get_sheet_names()
# ['Sheet'] -> default name prolly. 
# lets select this sheet and work with it

sheet = wb.get_sheet_by_name('Sheet')

sheet['A1'].value == None
# returns true

sheet['A1'] = 42
sheet['A2'] = 'Hello'

# save the sheet to some other directory. 

import os

os.chdir('c:\\Users\\UserName\\Documents')
wb.save()

# create another sheet
sheet2 = wb.create_sheet()
wb.get_sheet_names()

# would returns 2 sheets now
# [Sheet1, Sheet2]

sheet2.title = 'New Sheet Of People Name'
wb.get_sheet_names()

# now it will return sheet1 and New Sheet of People Name
wb.save('example.xlsx')
# workbook saved

wb.create_sheet(index=0, title='My Other Sheet')
wb.save('Other Sheet.xlsx')
# Now, this workbook would have first sheet named my other sheet
# as you can see the first para index=0, represents first sheet index







